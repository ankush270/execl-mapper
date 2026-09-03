# -*- coding: utf-8 -*-
"""
P&C Insurance Test Data Generator  v2.0
========================================
Generates 6 professionally designed Excel templates + matching JSON files.
Each sheet has its own unique color theme and a clean 2-column Label | Value layout.

Domains:
  1. Auto Insurance Policy Application         – Steel Blue
  2. Home Insurance Policy                     – Forest Green
  3. Commercial General Liability (CGL)        – Deep Burgundy
  4. Workers Compensation Report               – Dark Teal
  5. Claims FNOL (First Notice of Loss)        – Burnt Orange
  6. Reinsurance Treaty Summary               – Royal Purple

Excel templates  -> backend/templates/
JSON data files  -> backend/test_json/
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    sys.exit("openpyxl is not installed. Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR     = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
JSON_DIR     = BASE_DIR / "test_json"
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
JSON_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Color Palettes  (one per sheet)
# ---------------------------------------------------------------------------
THEMES = {
    "auto":   {"primary": "1B3A6B", "light": "D6E4F7", "accent": "2E75B6", "row_alt": "EBF3FB"},
    "home":   {"primary": "1E4D2B", "light": "D4EDDA", "accent": "28A745", "row_alt": "EBF5EE"},
    "cgl":    {"primary": "6B1A1A", "light": "F5D6D6", "accent": "C0392B", "row_alt": "FDF0F0"},
    "wc":     {"primary": "0D4D4D", "light": "CCE8E8", "accent": "17A2B8", "row_alt": "E8F6F8"},
    "fnol":   {"primary": "7D3200", "light": "FAE3CC", "accent": "E67E22", "row_alt": "FDF3E7"},
    "reins":  {"primary": "3D1080", "light": "E2D4F5", "accent": "8E44AD", "row_alt": "F5EFFE"},
}

# ---------------------------------------------------------------------------
# Generic style helpers
# ---------------------------------------------------------------------------
def _side():
    return Side(style="thin", color="AAAAAA")

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom(color="555555"):
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin",   color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def _no_border():
    return Border()

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", indent=1, wrap_text=wrap)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(size=10, bold=False, color="1A1A1A", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

# ---------------------------------------------------------------------------
# High-level cell writers
# ---------------------------------------------------------------------------
def write_title(ws, row, text, col_start, col_end, theme):
    """Large centered title spanning col_start:col_end."""
    ws.row_dimensions[row].height = 34
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font      = _font(size=15, bold=True, color="FFFFFF")
    c.fill      = _fill(theme["primary"])
    c.alignment = _center()
    c.border    = _border()


def write_subtitle(ws, row, text, col_start, col_end, theme):
    """Small meta-info subtitle row (company name, date, etc.)."""
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font      = _font(size=9, italic=True, color="FFFFFF")
    c.fill      = _fill(theme["accent"])
    c.alignment = _center()


def write_section(ws, row, text, col_start, col_end, theme):
    """Section heading — full width, colored background."""
    ws.row_dimensions[row].height = 20
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font      = _font(size=12, bold=True, color=theme["primary"])
    c.fill      = _fill(theme["light"])
    c.alignment = _left()
    c.border    = _thick_bottom(theme["primary"])


def write_field(ws, row, label, col_label, col_value, theme, alt=False):
    """One label + one empty value cell."""
    ws.row_dimensions[row].height = 18

    # Label
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.font      = _font(size=10, bold=True, color="2C2C2C")
    lc.fill      = _fill("F2F2F2") if not alt else _fill("EAEAEA")
    lc.alignment = _left()
    lc.border    = _border()

    # Value
    vc = ws.cell(row=row, column=col_value, value="")
    vc.font      = _font(size=10, color="1A1A1A")
    vc.fill      = _fill("FFFFFF") if not alt else _fill(theme["row_alt"])
    vc.alignment = _left()
    vc.border    = _border()


def write_col_headers(ws, row, headers, start_col, theme):
    """Table-style column headers (for multi-column tables)."""
    ws.row_dimensions[row].height = 20
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font      = _font(size=10, bold=True, color="FFFFFF")
        c.fill      = _fill(theme["accent"])
        c.alignment = _center()
        c.border    = _border()


def set_col_widths(ws, widths: dict):
    """widths = {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def blank_row(ws, row):
    ws.row_dimensions[row].height = 6


# ===========================================================================
# 1. AUTO INSURANCE POLICY APPLICATION  (Steel Blue)
# ===========================================================================
def create_auto_policy():
    t   = THEMES["auto"]
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Auto Policy Application"
    ws.sheet_view.showGridLines = False

    # Column layout: A=margin | B=label | C=value
    set_col_widths(ws, {"A": 2, "B": 30, "C": 42})
    COL_L, COL_V, SPAN_END = 2, 3, 3   # column B=2, C=3

    r = 1
    write_title(ws, r, "AUTO INSURANCE POLICY APPLICATION", COL_L, SPAN_END, t);    r += 1
    write_subtitle(ws, r, "Personal Lines Division  |  Policy Issuance Form",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    # -- Policyholder Information --
    write_section(ws, r, "  POLICYHOLDER INFORMATION", COL_L, SPAN_END, t);          r += 1
    ph_fields = [
        "Policy Number", "Policy Effective Date", "Policy Expiry Date",
        "First Name", "Last Name", "Date of Birth", "Gender",
        "Marital Status", "License Number", "License State",
        "Phone Number", "Email Address",
    ]
    for i, lbl in enumerate(ph_fields):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    # -- Address --
    write_section(ws, r, "  INSURED ADDRESS", COL_L, SPAN_END, t); r += 1
    for i, lbl in enumerate(["Street Address", "City", "State", "ZIP Code", "County"]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    # -- Vehicle --
    write_section(ws, r, "  VEHICLE INFORMATION", COL_L, SPAN_END, t); r += 1
    for i, lbl in enumerate([
        "VIN", "Make", "Model", "Year", "Color",
        "Vehicle Usage", "Annual Mileage", "Garaging ZIP"
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    # -- Coverage --
    write_section(ws, r, "  COVERAGE DETAILS", COL_L, SPAN_END, t); r += 1
    for i, lbl in enumerate([
        "Bodily Injury Liability", "Property Damage Liability",
        "Collision Deductible", "Comprehensive Deductible",
        "Uninsured Motorist", "Medical Payments",
        "Annual Premium", "Payment Frequency",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "01_auto_policy_application.xlsx")
    print("[OK] 01_auto_policy_application.xlsx")


# ===========================================================================
# 2. HOME INSURANCE POLICY  (Forest Green)
# ===========================================================================
def create_home_policy():
    t  = THEMES["home"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Home Insurance Policy"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 2, "B": 30, "C": 42})
    COL_L, COL_V, SPAN_END = 2, 3, 3

    r = 1
    write_title(ws, r, "HOMEOWNERS INSURANCE POLICY", COL_L, SPAN_END, t);           r += 1
    write_subtitle(ws, r, "Property & Casualty  |  Personal Lines  |  HO-3 Form",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    write_section(ws, r, "  POLICY INFORMATION", COL_L, SPAN_END, t);                r += 1
    for i, lbl in enumerate([
        "Policy Number", "Insurer Name", "Agent Name", "Agent Code",
        "Effective Date", "Expiry Date", "Policy Type",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  INSURED DETAILS", COL_L, SPAN_END, t);                   r += 1
    for i, lbl in enumerate([
        "Policyholder Name", "Co-Insured Name",
        "Date of Birth", "Phone Number", "Email Address",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  PROPERTY DETAILS", COL_L, SPAN_END, t);                  r += 1
    for i, lbl in enumerate([
        "Property Address", "City", "State", "ZIP Code",
        "Year Built", "Construction Type", "Roof Type",
        "Square Footage", "Number of Stories", "Occupancy Type",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  COVERAGE & LIMITS", COL_L, SPAN_END, t);                 r += 1
    for i, lbl in enumerate([
        "Dwelling Coverage", "Other Structures Coverage",
        "Personal Property Coverage", "Loss of Use Coverage",
        "Personal Liability Coverage", "Medical Payments Coverage",
        "All Perils Deductible", "Hurricane Deductible",
        "Annual Premium", "Flood Zone",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "02_home_insurance_policy.xlsx")
    print("[OK] 02_home_insurance_policy.xlsx")


# ===========================================================================
# 3. COMMERCIAL GENERAL LIABILITY (CGL)  (Deep Burgundy)
# ===========================================================================
def create_cgl_policy():
    t  = THEMES["cgl"]
    wb = Workbook()
    ws = wb.active
    ws.title = "CGL Policy"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 2, "B": 33, "C": 42})
    COL_L, COL_V, SPAN_END = 2, 3, 3

    r = 1
    write_title(ws, r, "COMMERCIAL GENERAL LIABILITY POLICY", COL_L, SPAN_END, t);   r += 1
    write_subtitle(ws, r, "Commercial Lines Division  |  Occurrence Form CG 00 01",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    write_section(ws, r, "  NAMED INSURED & POLICY INFO", COL_L, SPAN_END, t);       r += 1
    for i, lbl in enumerate([
        "Policy Number", "Named Insured", "DBA / Trade Name",
        "FEIN / Tax ID", "Business Type", "SIC Code",
        "Policy Effective Date", "Policy Expiry Date",
        "Carrier Name", "Underwriter Name",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  BUSINESS LOCATION", COL_L, SPAN_END, t);                 r += 1
    for i, lbl in enumerate([
        "Street Address", "City", "State", "ZIP Code",
        "Total Locations", "Description of Operations",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  COVERAGE LIMITS", COL_L, SPAN_END, t);                   r += 1
    for i, lbl in enumerate([
        "Each Occurrence Limit", "General Aggregate Limit",
        "Products Completed Ops Aggregate", "Personal Advertising Injury",
        "Medical Expense Limit", "Damage to Rented Premises",
        "Deductible Per Occurrence", "Retroactive Date",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  PREMIUM INFORMATION", COL_L, SPAN_END, t);               r += 1
    for i, lbl in enumerate([
        "Gross Annual Revenue", "Number of Employees",
        "Total Payroll", "Annual Premium", "Minimum Earned Premium",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "03_cgl_policy.xlsx")
    print("[OK] 03_cgl_policy.xlsx")


# ===========================================================================
# 4. WORKERS COMPENSATION  (Dark Teal)
# ===========================================================================
def create_workers_comp():
    t  = THEMES["wc"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Workers Compensation"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 2, "B": 33, "C": 42})
    COL_L, COL_V, SPAN_END = 2, 3, 3

    r = 1
    write_title(ws, r, "WORKERS COMPENSATION POLICY REPORT", COL_L, SPAN_END, t);    r += 1
    write_subtitle(ws, r, "Commercial Lines  |  Statutory Benefits  |  NCCI Form",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    write_section(ws, r, "  EMPLOYER INFORMATION", COL_L, SPAN_END, t);              r += 1
    for i, lbl in enumerate([
        "Policy Number", "Employer Name", "FEIN", "NAICS Code",
        "Industry Class", "Contact Name", "Contact Phone",
        "Effective Date", "Expiry Date", "State of Coverage",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  EMPLOYEE STATISTICS", COL_L, SPAN_END, t);               r += 1
    for i, lbl in enumerate([
        "Total Employees", "Full Time Employees", "Part Time Employees",
        "Total Annual Payroll", "Experience Modification Factor",
        "Class Code", "Class Description",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  CLAIMS HISTORY (PRIOR 3 YEARS)", COL_L, SPAN_END, t);   r += 1
    for i, lbl in enumerate([
        "Total Claims Count", "Total Incurred Losses",
        "Total Paid Losses", "Total Reserved Losses",
        "Lost Time Claims", "Medical Only Claims",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  PREMIUM CALCULATION", COL_L, SPAN_END, t);               r += 1
    for i, lbl in enumerate([
        "Manual Premium", "Experience Mod Adjusted Premium",
        "Schedule Credit Debit", "Net Premium",
        "Expense Constant", "Total Annual Premium",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "04_workers_compensation.xlsx")
    print("[OK] 04_workers_compensation.xlsx")


# ===========================================================================
# 5. CLAIMS FNOL  (Burnt Orange)
# ===========================================================================
def create_claims_fnol():
    t  = THEMES["fnol"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Claims FNOL"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 2, "B": 30, "C": 44})
    COL_L, COL_V, SPAN_END = 2, 3, 3

    r = 1
    write_title(ws, r, "FIRST NOTICE OF LOSS (FNOL) REPORT", COL_L, SPAN_END, t);   r += 1
    write_subtitle(ws, r, "Claims Management Division  |  Intake Form  |  Confidential",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    write_section(ws, r, "  CLAIM IDENTIFICATION", COL_L, SPAN_END, t);              r += 1
    for i, lbl in enumerate([
        "Claim Number", "Policy Number", "Date of Loss",
        "Date Reported", "Claim Status", "Line of Business",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  CLAIMANT INFORMATION", COL_L, SPAN_END, t);              r += 1
    for i, lbl in enumerate([
        "Claimant First Name", "Claimant Last Name",
        "Claimant Date of Birth", "Claimant Phone",
        "Claimant Email", "Relationship to Insured",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  LOSS DETAILS", COL_L, SPAN_END, t);                      r += 1
    for i, lbl in enumerate([
        "Loss Location Address", "Loss City", "Loss State",
        "Cause of Loss", "Description of Loss",
        "Police Report Number", "Weather Conditions",
        "Injuries Reported", "Property Damage Reported",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  ADJUSTER ASSIGNMENT", COL_L, SPAN_END, t);               r += 1
    for i, lbl in enumerate([
        "Adjuster Name", "Adjuster ID", "Adjuster Phone",
        "Supervisor Name", "Estimated Loss Amount",
        "Reserve Amount", "Coverage Verified",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "05_claims_fnol.xlsx")
    print("[OK] 05_claims_fnol.xlsx")


# ===========================================================================
# 6. REINSURANCE TREATY  (Royal Purple)
# ===========================================================================
def create_reinsurance_treaty():
    t  = THEMES["reins"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Reinsurance Treaty"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, {"A": 2, "B": 33, "C": 44})
    COL_L, COL_V, SPAN_END = 2, 3, 3

    r = 1
    write_title(ws, r, "REINSURANCE TREATY SUMMARY SHEET", COL_L, SPAN_END, t);      r += 1
    write_subtitle(ws, r, "Reinsurance Division  |  Treaty Administration  |  Confidential",
                   COL_L, SPAN_END, t);                                               r += 1
    blank_row(ws, r);                                                                 r += 1

    write_section(ws, r, "  TREATY IDENTIFICATION", COL_L, SPAN_END, t);             r += 1
    for i, lbl in enumerate([
        "Treaty Number", "Treaty Name", "Treaty Type",
        "Cedant Name", "Cedant Country", "Reinsurer Name",
        "Broker Name", "Treaty Effective Date", "Treaty Expiry Date",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  TREATY STRUCTURE", COL_L, SPAN_END, t);                  r += 1
    for i, lbl in enumerate([
        "Line of Business", "Territory",
        "Retention Amount", "Cession Limit",
        "Cession Percentage", "Reinsurance Premium Rate",
        "Commission Rate", "Loss Ratio Trigger",
        "Aggregate Limit", "Occurrence Limit",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    blank_row(ws, r); r += 1

    write_section(ws, r, "  FINANCIAL SUMMARY", COL_L, SPAN_END, t);                 r += 1
    for i, lbl in enumerate([
        "Ceded Premium", "Ceded Losses", "Ceded Loss Ratio",
        "Net Premium Retained", "Reinsurance Commission Earned",
        "Profit Commission", "Currency", "Exchange Rate",
    ]):
        write_field(ws, r, lbl, COL_L, COL_V, t, alt=(i % 2 == 1)); r += 1

    _add_footer(ws, r + 1, COL_L, SPAN_END, t)
    wb.save(TEMPLATE_DIR / "06_reinsurance_treaty.xlsx")
    print("[OK] 06_reinsurance_treaty.xlsx")


# ---------------------------------------------------------------------------
# Footer helper
# ---------------------------------------------------------------------------
def _add_footer(ws, row, col_start, col_end, theme):
    blank_row(ws, row)
    row += 1
    ws.row_dimensions[row].height = 14
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start,
                value="CONFIDENTIAL  |  FOR INTERNAL USE ONLY  |  Generated by Excel Mapping Tool")
    c.font      = _font(size=8, italic=True, color="888888")
    c.alignment = _center()
    c.fill      = _fill("F7F7F7")


# ===========================================================================
# JSON Data Files
# ===========================================================================
def _save_json(filename, data):
    path = JSON_DIR / filename
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
    print(f"[OK] {filename}")


def create_auto_policy_json():
    _save_json("01_auto_policy.json", {
        "policy_number": "AUTO-2024-001234",
        "policy_effective_date": "2024-01-01",
        "policy_expiry_date": "2025-01-01",
        "first_name": "Rajesh",
        "last_name": "Sharma",
        "date_of_birth": "1985-06-15",
        "gender": "Male",
        "marital_status": "Married",
        "license_number": "DL-TX-987654",
        "license_state": "TX",
        "phone_number": "+1-512-555-0199",
        "email_address": "rajesh.sharma@email.com",
        "street_address": "4521 Oak Creek Dr",
        "city": "Austin",
        "state": "TX",
        "zip_code": "78701",
        "county": "Travis",
        "vin": "1HGBH41JXMN109186",
        "make": "Honda",
        "model": "Accord",
        "year": 2021,
        "color": "Silver",
        "vehicle_usage": "Commute",
        "annual_mileage": 12000,
        "garaging_zip": "78701",
        "bodily_injury_liability": "100/300",
        "property_damage_liability": "100,000",
        "collision_deductible": 500,
        "comprehensive_deductible": 250,
        "uninsured_motorist": "100/300",
        "medical_payments": 5000,
        "annual_premium": 1248.00,
        "payment_frequency": "Monthly",
    })


def create_home_policy_json():
    _save_json("02_home_insurance_policy.json", {
        "policy_number": "HO-2024-055891",
        "insurer_name": "Liberty Shield Insurance Co.",
        "agent_name": "Priya Mehta",
        "agent_code": "AGT-4421",
        "effective_date": "2024-03-01",
        "expiry_date": "2025-03-01",
        "policy_type": "HO-3 Special Form",
        "policyholder_name": "Ankush Verma",
        "co_insured_name": "Sneha Verma",
        "date_of_birth": "1980-11-22",
        "phone_number": "+1-469-555-0177",
        "email_address": "ankush.verma@email.com",
        "property_address": "789 Maple Lane",
        "city": "Dallas",
        "state": "TX",
        "zip_code": "75201",
        "year_built": 2005,
        "construction_type": "Frame",
        "roof_type": "Asphalt Shingle",
        "square_footage": 2400,
        "number_of_stories": 2,
        "occupancy_type": "Owner-Occupied",
        "dwelling_coverage": 350000,
        "other_structures_coverage": 35000,
        "personal_property_coverage": 175000,
        "loss_of_use_coverage": 70000,
        "personal_liability_coverage": 300000,
        "medical_payments_coverage": 5000,
        "all_perils_deductible": 1000,
        "hurricane_deductible": 5000,
        "annual_premium": 2184.50,
        "flood_zone": "Zone X",
    })


def create_cgl_json():
    _save_json("03_cgl_policy.json", {
        "policy_number": "CGL-2024-789012",
        "named_insured": "Sunrise Tech Solutions LLC",
        "dba_trade_name": "Sunrise Tech",
        "fein_tax_id": "47-1234567",
        "business_type": "Limited Liability Company",
        "sic_code": "7372",
        "policy_effective_date": "2024-07-01",
        "policy_expiry_date": "2025-07-01",
        "carrier_name": "Pinnacle Insurance Group",
        "underwriter_name": "James O'Brien",
        "street_address": "2200 Commerce Pkwy",
        "city": "Houston",
        "state": "TX",
        "zip_code": "77002",
        "total_locations": 3,
        "description_of_operations": "Software development and IT consulting services",
        "each_occurrence_limit": 1000000,
        "general_aggregate_limit": 2000000,
        "products_completed_ops_aggregate": 2000000,
        "personal_advertising_injury": 1000000,
        "medical_expense_limit": 10000,
        "damage_to_rented_premises": 300000,
        "deductible_per_occurrence": 5000,
        "retroactive_date": "2020-07-01",
        "gross_annual_revenue": 4500000,
        "number_of_employees": 45,
        "total_payroll": 3150000,
        "annual_premium": 18750.00,
        "minimum_earned_premium": 9375.00,
    })


def create_workers_comp_json():
    _save_json("04_workers_compensation.json", {
        "policy_number": "WC-2024-334455",
        "employer_name": "BuildRight Construction Inc.",
        "fein": "83-7654321",
        "naics_code": "236220",
        "industry_class": "Commercial Building Construction",
        "contact_name": "Maria Gonzalez",
        "contact_phone": "+1-713-555-0233",
        "effective_date": "2024-04-01",
        "expiry_date": "2025-04-01",
        "state_of_coverage": "TX",
        "total_employees": 120,
        "full_time_employees": 95,
        "part_time_employees": 25,
        "total_annual_payroll": 5800000,
        "experience_modification_factor": 0.92,
        "class_code": "5645",
        "class_description": "Carpentry - Residential Construction",
        "total_claims_count": 8,
        "total_incurred_losses": 142000,
        "total_paid_losses": 98000,
        "total_reserved_losses": 44000,
        "lost_time_claims": 3,
        "medical_only_claims": 5,
        "manual_premium": 87000,
        "experience_mod_adjusted_premium": 80040,
        "schedule_credit_debit": -4002,
        "net_premium": 76038,
        "expense_constant": 350,
        "total_annual_premium": 76388.00,
    })


def create_claims_fnol_json():
    _save_json("05_claims_fnol.json", {
        "claim_number": "CLM-2024-998877",
        "policy_number": "AUTO-2024-001234",
        "date_of_loss": "2024-08-14",
        "date_reported": "2024-08-14",
        "claim_status": "Open - Under Investigation",
        "line_of_business": "Personal Auto",
        "claimant_first_name": "Sunita",
        "claimant_last_name": "Patel",
        "claimant_date_of_birth": "1990-03-10",
        "claimant_phone": "+1-214-555-0155",
        "claimant_email": "sunita.patel@email.com",
        "relationship_to_insured": "Spouse",
        "loss_location_address": "I-35 & Hwy 183 Intersection",
        "loss_city": "Austin",
        "loss_state": "TX",
        "cause_of_loss": "Rear-end Collision",
        "description_of_loss": "Insured vehicle was rear-ended at a traffic stop. Airbags deployed.",
        "police_report_number": "APD-2024-112233",
        "weather_conditions": "Clear / Dry",
        "injuries_reported": True,
        "property_damage_reported": True,
        "adjuster_name": "Carlos Rivera",
        "adjuster_id": "ADJ-7788",
        "adjuster_phone": "+1-800-555-0301",
        "supervisor_name": "Linda Nakamura",
        "estimated_loss_amount": 22500,
        "reserve_amount": 25000,
        "coverage_verified": True,
    })


def create_reinsurance_json():
    _save_json("06_reinsurance_treaty.json", {
        "treaty_number": "TRT-2024-RE-4421",
        "treaty_name": "Commercial Lines Quota Share Treaty 2024",
        "treaty_type": "Quota Share",
        "cedant_name": "Pinnacle Insurance Group",
        "cedant_country": "United States",
        "reinsurer_name": "Munich Re America",
        "broker_name": "Aon Benfield",
        "treaty_effective_date": "2024-01-01",
        "treaty_expiry_date": "2024-12-31",
        "line_of_business": "Commercial Property",
        "territory": "United States & Canada",
        "retention_amount": 2000000,
        "cession_limit": 8000000,
        "cession_percentage": 60.0,
        "reinsurance_premium_rate": 0.18,
        "commission_rate": 0.32,
        "loss_ratio_trigger": 0.70,
        "aggregate_limit": 50000000,
        "occurrence_limit": 10000000,
        "ceded_premium": 12600000,
        "ceded_losses": 7840000,
        "ceded_loss_ratio": 62.22,
        "net_premium_retained": 8400000,
        "reinsurance_commission_earned": 4032000,
        "profit_commission": 287500,
        "currency": "USD",
        "exchange_rate": 1.0,
    })


# ===========================================================================
# Entry Point
# ===========================================================================
if __name__ == "__main__":
    print("\n[*] Generating P&C Insurance Excel Templates...\n")
    create_auto_policy()
    create_home_policy()
    create_cgl_policy()
    create_workers_comp()
    create_claims_fnol()
    create_reinsurance_treaty()

    print("\n[*] Generating Matching JSON Data Files...\n")
    create_auto_policy_json()
    create_home_policy_json()
    create_cgl_json()
    create_workers_comp_json()
    create_claims_fnol_json()
    create_reinsurance_json()

    print("\n[DONE]")
    print(f"   Excel templates -> {TEMPLATE_DIR}")
    print(f"   JSON data files -> {JSON_DIR}\n")
