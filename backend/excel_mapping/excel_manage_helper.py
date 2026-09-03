from __future__ import annotations
import copy
import json
import re
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Any
from bson import ObjectId
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from core_utils.config import Config
from core_utils.logging import logger
from openpyxl.utils.cell import coordinate_from_string

class ExcelManageHelper:
    
    def to_object_id(self, value: str | ObjectId) -> ObjectId:
        if isinstance(value, ObjectId):
            return value
        if value and ObjectId.is_valid(value):
            return ObjectId(value)
        raise ValueError("Invalid ObjectId.")

    @staticmethod
    def normalize_name(value: Any) -> str:
        if not value:
            return ""
        cleaned = re.sub(r"[^a-z0-9\s]", " ", str(value).lower().replace("_", " ").replace("-", " "))
        res = re.sub(r"\s+", " ", cleaned).strip()
        return res if res else str(value).lower().strip()

    @staticmethod
    def name_tokens(value: Any) -> set[str]:
        return set(ExcelManageHelper.normalize_name(value).split())

    def clean_value(self, value: Any) -> Any:
        return value.isoformat() if isinstance(value, (datetime, date)) else value

    def join_json_path(self, path: str, key: str) -> str:
        key = str(key).replace("\\", "\\\\").replace("'", "\\'")
        return f"{path}['{key}']"

    def safe_template_path(self, file_name: str) -> Path:
        if not file_name:
            raise ValueError("template_file is required.")

        requested = Path(file_name)

        if (requested.name != file_name or requested.suffix.lower() not in {".xlsx", ".xlsm"}):
            raise ValueError("Invalid Excel template.")

        template_folder = Path(Config.TEMPLATE_FOLDER).resolve()
        template_path = (template_folder / requested.name).resolve()

        if not template_path.exists():
            raise ValueError("Excel template not found.")
        return template_path

    def analyze_template_workbook(self, workbook, *, file_name: str, only_sheet: str | None = None) -> dict[str, Any]:
        analyzer = _WorkbookAnalyzerAdapter(workbook)
        schema = analyzer.analyze()
        if only_sheet:
            schema["sheets"] = [s for s in schema["sheets"] if s["name"] == only_sheet]
            if not schema["sheets"]:
                raise ValueError(f"Sheet '{only_sheet}' does not exist.")
        return schema

    def analyze_json(self, data: Any) -> dict[str, Any]:
        fields: list[dict[str, Any]] = []
        arrays: list[dict[str, Any]] = []
        self._walk_json(value=data, path="$", parent_path="$", parent_name=None, fields=fields, arrays=arrays)
        return {"root_type": self._get_type(data), "fields": fields, "arrays": arrays}

    _MAX_JSON_DEPTH = 50

    def _create_field_entry(self, key: Any, child: Any, child_path: str, parent_path: str, parent_name: str | None) -> dict[str, Any]:
        return {
            "name": str(key),
            "normalized_name": self.normalize_name(key),
            "type": self._get_type(child),
            "path": child_path,
            "parent_path": parent_path,
            "parent_name": str(parent_name) if parent_name is not None else None,
            "value": self.clean_value(child),
            "array_depth": child_path.count("[*]"),
            "tokens": list(self.name_tokens(key)),
        }

    def _walk_json(
        self,
        value: Any,
        path: str,
        parent_path: str,
        parent_name: str | None,
        fields: list[dict[str, Any]],
        arrays: list[dict[str, Any]],
        _depth: int = 0,
    ):
        if _depth > self._MAX_JSON_DEPTH:
            return
        if isinstance(value, dict):
            for key, child in value.items():
                child_path = self.join_json_path(path, key)
                if isinstance(child, (dict, list)):
                    self._walk_json(child, child_path, path, key, fields, arrays, _depth=_depth + 1)
                else:
                    fields.append(self._create_field_entry(key, child, child_path, path, parent_name))

        elif isinstance(value, list):
            array_entry = {"path": path, "parent_path": parent_path, "parent_name": parent_name, "length": len(value)}

            if not any(a["path"] == path for a in arrays):
                arrays.append(array_entry)

            if value:
                representative = value[0]
                if isinstance(representative, dict):
                    merged_keys: dict[str, Any] = {}
                    for item in value[:20]:
                        if isinstance(item, dict):
                            for k, v in item.items():
                                if k not in merged_keys or merged_keys[k] is None:
                                    merged_keys[k] = v

                    item_path = f"{path}[*]"
                    for key, child in merged_keys.items():
                        child_path = self.join_json_path(item_path, key)
                        if isinstance(child, (dict, list)):
                            self._walk_json(child, child_path, f"{path}[*]", key, fields, arrays, _depth=_depth + 1)
                        else:
                            fields.append(self._create_field_entry(key, child, child_path, f"{path}[*]", parent_name))

                elif isinstance(representative, list):
                    self._walk_json(representative, f"{path}[*]", f"{path}[*]", parent_name, fields, arrays, _depth=_depth + 1)

    def _get_type(self, value: Any) -> str:
        type_map = {type(None): "null", bool: "boolean", int: "integer", float: "number", str: "string", list: "array", dict: "object"}
        return type_map.get(type(value), "unknown")

    def build_mappings(self, *, json_schema: dict[str, Any], excel_sheet_schema: dict[str, Any]) -> list[dict[str, Any]]:
        json_fields = json_schema.get("fields", [])
        excel_fields = self._flatten_excel_fields(excel_sheet_schema)
        mappings: list[dict[str, Any]] = []
        used_json_paths: set[str] = set()

        for excel_field in excel_fields:
            best = None
            best_score = 0.0

            for json_field in json_fields:
                if json_field["path"] in used_json_paths:
                    continue
                score = self._mapping_score(excel_field, json_field)
                if score > best_score:
                    best_score = score
                    best = json_field
            if best is None or best_score < 0.50:
                continue

            used_json_paths.add(best["path"])

            mapping = {
                "excel": {
                    "sheet": excel_sheet_schema["name"],
                    "label": excel_field["name"],
                    "label_cell": excel_field["label_cell"],
                    "value_cell": excel_field["value_cell"],
                    "row": excel_field["row"],
                    "column": excel_field["column"],
                    "level": excel_field.get("level"),
                    "section_path": excel_field.get("section_path", []),
                },
                "json": {
                    "field": best["name"],
                    "path": best["path"],
                    "type": best["type"],
                    "array_depth": best.get("array_depth", 0),
                },
                "header_type": best["type"],
                "header_name": excel_field["name"],
                "header_value": best.get("value"),
                "excel_column": excel_field["column"],
                "json_path": best["path"],
                "confidence": round(best_score, 4),
                "is_repeating": "[*]" in best["path"],
            }
            mappings.append(mapping)
        mappings.sort(key=lambda m: (m["excel"]["row"], m["excel"]["column"]))
        return mappings

    def _format_excel_field(self, field: dict[str, Any], section_path: list[str]) -> dict[str, Any]:
        val_cell = field.get("value_cell")
        col = coordinate_from_string(val_cell)[0] if val_cell else None
        return {
            "name": field.get("name", ""),
            "normalized_name": self.normalize_name(field.get("name")),
            "row": field.get("row"),
            "label_cell": field.get("label_cell"),
            "value_cell": val_cell,
            "column": col,
            "level": field.get("level"),
            "section_path": section_path,
            "expected_type": field.get("expected_type"),
        }

    def _flatten_excel_fields(self, sheet_schema: dict[str, Any]) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []

        def visit_section(section: dict[str, Any], parent_names: list[str]):
            current_path = parent_names + [section["name"]]
            for field in section.get("fields", []):
                result.append(self._format_excel_field(field, current_path))
            for child in section.get("children", []):
                visit_section(child, current_path)

        for section in sheet_schema.get("sections", []):
            visit_section(section, [])

        for field in sheet_schema.get("orphan_fields", []):
            result.append(self._format_excel_field(field, []))

        return [field for field in result if field["name"] and field["value_cell"]]

    def _mapping_score(self, excel_field: dict[str, Any], json_field: dict[str, Any]) -> float:
        excel_name = excel_field["normalized_name"]
        json_name = json_field["normalized_name"]

        if not excel_name or not json_name:
            return 0.0

        score = 0.0

        if excel_name == json_name:
            score += 0.70
        else:
            excel_tokens = self.name_tokens(excel_name)
            json_tokens = self.name_tokens(json_name)

            if excel_tokens and json_tokens:
                intersection = len(excel_tokens & json_tokens)
                union = len(excel_tokens | json_tokens)
                similarity = intersection / union
                score += 0.45 * similarity
        excel_context = {self.normalize_name(x) for x in excel_field.get("section_path", []) if x}
        json_context = set()
        parent = json_field.get("parent_name")
        if parent:
            json_context.add(self.normalize_name(parent))
        path_parts = re.findall(r"\['((?:\\.|[^'])*)'\]", json_field["path"])
        json_context.update(self.normalize_name(x) for x in path_parts[:-1] if x)

        if excel_context and json_context:
            overlap = len(excel_context & json_context)
            if overlap:
                score += min(0.20, overlap * 0.10)

        expected = excel_field.get("expected_type")
        actual = json_field.get("type")

        if expected and actual:
            if self._types_compatible(expected, actual):
                score += 0.10
            else:
                score -= 0.05

        if json_field.get("array_depth", 0) > 0:
            score += 0.02

        return max(0.0, min(score, 1.0))

    def _types_compatible(self, expected: str, actual: str) -> bool:
        aliases = {"int": "integer", "float": "number", "double": "number", "bool": "boolean"}
        exp, act = aliases.get(expected, expected), aliases.get(actual, actual)
        return exp == act or (exp == "number" and act == "integer") or (exp == "string" and act in {"string", "date", "integer", "number", "boolean"})

    def get_json_value(self, data: Any, path: str) -> list[Any]:
        try:
            tokens = self._parse_json_path(path)
            values = [data]
            for token in tokens:
                next_values = []
                if token == "[*]":
                    for value in values:
                        if isinstance(value, list):
                            next_values.extend(value)
                else:
                    for value in values:
                        if isinstance(value, dict) and token in value:
                            next_values.append(value[token])
                values = next_values
                if not values:
                    break
            return values

        except Exception as exc:
            logger.error(f"Unable to read JSON path '{path}': {exc}")
            raise ValueError(f"Unable to read JSON path '{path}': {exc}") from exc

    def _parse_json_path(self, path: str) -> list[str]:
        if not path or path == "$":
            return []
        tokens = []
        pattern = re.compile(r"\['((?:\\.|[^'])*)'\]|\[\*\]")
        position = 1 if path.startswith("$") else 0

        while position < len(path):
            if path[position] == ".":
                position += 1
                continue
            match = pattern.match(path, position)
            if not match:
                raise ValueError(f"Unsupported JSON path syntax near: {path[position:]}")

            if match.group(0) == "[*]":
                tokens.append("[*]")
            else:
                key = match.group(1)
                key = key.replace("\\'", "'").replace("\\\\", "\\")
                tokens.append(key)
            position = match.end()
        return tokens

    def populate_template(self, *, workbook, sheet, json_data, mappings: list[dict[str, Any]]) -> dict[str, Any]:
        scalar_count = 0
        repeating_count = 0
        empty_count = 0
        truncated_count = 0
        repeating_groups = defaultdict(list)
        scalar_mappings = []

        for mapping in mappings:
            if mapping.get("is_repeating"):
                path = mapping["json_path"]
                marker = path.find("[*]")
                array_path = path if marker == -1 else path[: marker + 3]
                repeating_groups[array_path].append(mapping)
            else:
                scalar_mappings.append(mapping)

        occupied_rows: set[int] = set()
        for m in scalar_mappings:
            r = m["excel"].get("row")
            if r:
                occupied_rows.add(r)

        group_base_rows: dict[str, int] = {}
        for array_path, group in repeating_groups.items():
            base_row = min(m["excel"]["row"] for m in group)
            group_base_rows[array_path] = base_row
            for m in group:
                occupied_rows.add(m["excel"]["row"])

        for mapping in scalar_mappings:
            values = self.get_json_value(json_data, mapping["json_path"])
            value = values[0] if values else None
            cell = mapping["excel"]["value_cell"]

            if not cell:
                continue

            sheet[cell] = self.clean_value(value)

            if value is None:
                empty_count += 1
            else:
                scalar_count += 1

        for array_path, group in repeating_groups.items():
            values = self.get_json_value(json_data, array_path.replace("[*]", ""))
            rows = values[0] if values and isinstance(values[0], list) else []

            if not rows:
                for mapping in group:
                    sheet[mapping["excel"]["value_cell"]] = None
                    empty_count += 1
                continue

            base_row = group_base_rows[array_path]
            group_template_rows = {m["excel"]["row"] for m in group}

            safe_limit = None
            for r in sorted(occupied_rows):
                if r > base_row and r not in group_template_rows:
                    safe_limit = r
                    break

            max_items = (safe_limit - base_row) if safe_limit else len(rows)
            safe_rows = rows[:max_items]

            if len(rows) > max_items:
                truncated_count += len(rows) - max_items
                logger.warning(
                    f"Array '{array_path}' truncated: {len(rows)} items but only "
                    f"{max_items} rows available before next section (row {safe_limit})."
                )

            for index, record in enumerate(safe_rows):
                target_row = base_row + index
                for mapping in group:
                    json_path = mapping["json_path"]
                    relative_path = json_path[len(array_path) :].lstrip(".")
                    values = self.get_json_value(record, "$" + relative_path) if relative_path else [record]
                    value = values[0] if values else None
                    source_cell = mapping["excel"]["value_cell"]
                    column = coordinate_from_string(source_cell)[0]
                    target_cell = f"{column}{target_row}"

                    if target_row != mapping["excel"]["row"]:
                        source = sheet[source_cell]
                        target = sheet[target_cell]
                        if not isinstance(target, MergedCell) and not isinstance(source, MergedCell):
                            target._style = copy.copy(source._style)

                    sheet[target_cell] = self.clean_value(value)

                    if value is None:
                        empty_count += 1
                    else:
                        repeating_count += 1
        return {
            "scalar_values_written": scalar_count,
            "repeating_values_written": repeating_count,
            "empty_values": empty_count,
            "total_mappings": len(mappings),
            "truncated_values": truncated_count,
        }

class _WorkbookAnalyzerAdapter:
    MAX_SCAN_ROWS = 10_000
    MAX_SCAN_COLS = 200

    def __init__(self, workbook):
        self.workbook = workbook

    def analyze(self):
        sheets = []
        for ws in self.workbook.worksheets:
            if ws.sheet_state in ("hidden", "veryHidden"):
                continue
            rows, column_scores, last_populated_row = self._analyze_rows(ws)
            if not rows:
                sheets.append(
                    {
                        "name": ws.title,
                        "max_row": ws.max_row,
                        "max_column": ws.max_column,
                        "value_column": None,
                        "rows": [],
                        "sections": [],
                        "root": None,
                        "orphan_fields": [],
                        "needs_review": [],
                    }
                )
                continue

            font_levels = self._discover_font_levels(rows)
            needs_review = self._classify_rows(rows, font_levels)
            value_column = self._detect_value_column(ws, column_scores)
            sections, orphan_fields = self._build_hierarchy(rows, value_column, last_populated_row)

            sheets.append(
                {
                    "name": ws.title,
                    "max_row": last_populated_row,
                    "max_column": ws.max_column,
                    "value_column": value_column,
                    "rows": rows,
                    "sections": sections,
                    "root": sections[0] if sections else None,
                    "orphan_fields": orphan_fields,
                    "needs_review": needs_review,
                }
            )

        return {"file_name": "workbook", "sheets": sheets}

    def _analyze_rows(self, ws):
        results = []
        column_scores = defaultdict(float)
        max_col = min(ws.max_column or 1, self.MAX_SCAN_COLS)
        max_row = min(ws.max_row or 1, self.MAX_SCAN_ROWS)
        last_populated_row = 0

        for row_number in range(1, max_row + 1):
            if row_number in ws.row_dimensions and ws.row_dimensions[row_number].hidden:
                continue
            cells = [ws.cell(row_number, c) for c in range(1, max_col + 1)]
            if all(self._is_empty(cell) for cell in cells):
                continue

            last_populated_row = row_number
            results.append(self._build_row_analysis(row_number, cells))

            for cell in cells:
                if cell.column >= 2:
                    if not self._is_empty(cell):
                        column_scores[cell.column] += 2.0
                    elif cell.has_style:
                        column_scores[cell.column] += 0.25

        return results, column_scores, last_populated_row or max_row

    def _build_row_analysis(self, row_number, cells):
        label_cell = self._find_label_cell(cells)
        value_cell = self._find_value_cell(cells, label_cell)
        is_anchor, spans_cols = self._get_merge_info(label_cell) if label_cell else (False, False)

        return {
            "row": row_number,
            "label_cell": (label_cell.coordinate if label_cell else None),
            "label": (self._clean_label(label_cell.value) if label_cell else None),
            "value_cells": ([value_cell.coordinate] if value_cell else []),
            "has_data_value": bool(value_cell and not self._is_empty(value_cell)),
            "is_merged_header": is_anchor and spans_cols,
            "font_size": (label_cell.font.sz if label_cell else None),
            "bold": bool(label_cell.font.bold) if label_cell else False,
            "fill_type": (label_cell.fill.fill_type if label_cell else None),
            "fill_color": self._get_fill_color(label_cell) if label_cell else None,
            "indent": (label_cell.alignment.indent or 0) if label_cell else 0,
            "row_type": "unknown",
            "level": None,
            "confidence": 0.0,
            "span": 1,
        }

    def _find_label_cell(self, cells):
        return next((c for c in cells if not isinstance(c, MergedCell) and not self._is_empty(c)), None)

    def _find_value_cell(self, cells, label_cell):
        """Find value cell to the RIGHT of label. Left-of-label layouts are not supported."""
        if label_cell is None:
            return None
        styled_cell = None
        for cell in cells:
            if cell.column <= label_cell.column:
                continue
            if isinstance(cell, MergedCell):
                continue
            if self._is_formula(cell):
                continue
            if not self._is_empty(cell):
                return cell
            if cell.has_style and styled_cell is None:
                styled_cell = cell
        return styled_cell

    def _discover_font_levels(self, rows):
        sizes = sorted({r["font_size"] for r in rows if r.get("font_size") is not None}, reverse=True)
        return {size: level for level, size in enumerate(sizes)}

    def _classify_rows(self, rows, font_levels):
        needs_review = []
        if not rows:
            return needs_review
        has_variation = len(set(font_levels.values())) > 1
        root_assigned = False

        for row in rows:
            field_score = (0.40 * bool(row.get("has_data_value")) + 0.10 * (row["indent"] > 0) + 0.10 * (not row["bold"]))
            header_score = (
                0.20 * bool(row["bold"])
                + 0.20 * row["is_merged_header"]
                + 0.15 * (row["fill_type"] == "solid")
                + 0.25 * (has_variation and font_levels.get(row["font_size"]) == 0)
            )

            if row.get("has_data_value"):
                is_field = True
            else:
                is_field = field_score >= header_score
            confidence = field_score if is_field else header_score

            if confidence < 0.20:
                row["row_type"] = "uncertain"
                row["confidence"] = confidence
                needs_review.append(row["row"])
                continue

            if is_field:
                row["row_type"] = "field"
                row["level"] = 2
                row["confidence"] = confidence
                continue

            if has_variation:
                level = font_levels.get(row["font_size"], 1)
            elif not root_assigned:
                level = 0
            elif row.get("fill_type") == "solid" or row.get("is_merged_header"):
                level = 1
            else:
                level = max(1, int(row.get("indent", 0)) + 1)

            if level == 0:
                root_assigned = True

            row["row_type"] = "root_header" if level == 0 else "section"
            row["level"] = level
            row["confidence"] = confidence

            if confidence < 0.50:
                needs_review.append(row["row"])

        return needs_review

    def _detect_value_column(self, ws, column_scores):
        return get_column_letter(max(column_scores, key=column_scores.get)) if ws.max_column >= 2 and column_scores else None

    def _build_hierarchy(self, rows, value_column, max_row):
        sections = []
        orphan_fields = []
        stack = []

        for row in rows:
            if row["row_type"] == "uncertain":
                continue

            if row["row_type"] in {"root_header", "section"}:
                section = {
                    "name": row["label"] or "",
                    "normalized_name": ExcelManageHelper.normalize_name(row["label"]),
                    "row": row["row"],
                    "level": row["level"],
                    "label_cell": row["label_cell"] or "",
                    "start_row": row["row"],
                    "end_row": None,
                    "fields": [],
                    "children": [],
                    "json_path": None,
                    "confidence": row["confidence"],
                }

                while stack and stack[-1][0] >= row["level"]:
                    stack.pop()

                if stack:
                    stack[-1][1]["children"].append(section)
                else:
                    sections.append(section)
                stack.append((row["level"], section))

            elif row["row_type"] == "field":
                value_cell = (
                    row["value_cells"][0]
                    if row["value_cells"]
                    else (f"{value_column}{row['row']}" if value_column else "")
                )

                field = {
                    "name": row["label"] or "",
                    "normalized_name": ExcelManageHelper.normalize_name(row["label"]),
                    "row": row["row"],
                    "label_cell": row["label_cell"] or "",
                    "value_cell": value_cell,
                    "level": 2,
                    "json_path": None,
                    "expected_type": None,
                    "confidence": row["confidence"],
                    "metadata": {},
                }

                if stack:
                    stack[-1][1]["fields"].append(field)
                else:
                    orphan_fields.append(field)

        self._close_section_ranges(sections, max_row)

        return sections, orphan_fields

    def _close_section_ranges(self, sections: list[dict[str, Any]], max_row: int):
        for index, section in enumerate(sections):
            section["end_row"] = (
                sections[index + 1]["row"] - 1 if index + 1 < len(sections) else max_row
            )
            children = section.get("children", [])
            if children:
                self._close_section_ranges(children, section["end_row"])

    def _is_empty(self, cell):
        return isinstance(cell, MergedCell) or cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())

    def _get_merge_info(self, cell):
        if not isinstance(cell, MergedCell):
            for r in cell.parent.merged_cells.ranges:
                if r.min_row == cell.row and r.min_col == cell.column:
                    return True, (r.max_col > r.min_col)
        return False, False

    def _is_formula(self, cell):
        return not isinstance(cell, MergedCell) and isinstance(cell.value, str) and cell.value.strip().startswith("=")

    def _get_fill_color(self, cell):
        try:
            if cell.fill and cell.fill.fill_type == "solid" and (c := cell.fill.fgColor):
                return str(c.indexed) if c.type == "indexed" else f"theme:{c.theme}" if c.type == "theme" else c.rgb if c.type == "rgb" else None
        except (AttributeError, TypeError):
            pass
        return None

    def _clean_label(self, value):
        return " ".join(str(value).split()) if value is not None else ""
