from __future__ import annotations
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
import json
from bson import ObjectId
from datetime import datetime, timezone
from openpyxl import load_workbook
from core_utils.config import Config
from core_utils.db import Database
from core_utils.logging import logger
from excel_mapping.excel_mapping_helper import ExcelManageHelper

class ExcelManageComponent:

    def __init__(self) -> None:
        self.db = Database()
        self.helper = ExcelManageHelper()

    def upload_json(self, file):
        if not file or not file.filename:
            raise ValueError("JSON file is required.")
        if Path(file.filename).suffix.lower() != ".json":
            raise ValueError("Only JSON files are allowed.")

        try:
            json_data = json.load(file)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid JSON file: {exc.msg} "
                             f"(line {exc.lineno}, column {exc.colno}).") from exc

        document = {
            "file_name": Path(file.filename).name,
            "json_data": json_data,
            "created_at": datetime.now(timezone.utc),}

        result = self.db.json_collection.insert_one(document)
        logger.info(f"JSON uploaded: {result.inserted_id}")
        return {"document_id": str(result.inserted_id),"file_name": document["file_name"],}

    def get_all_json(self):
        documents = self.db.json_collection.find({},{"json_data": 0},).sort("created_at", -1)
        return [{
                "document_id": str(document["_id"]),
                "file_name": document["file_name"],
                "created_at": document["created_at"],}for document in documents]

    def get_json(self, document_id: str):
        object_id = self.helper.to_object_id(document_id)
        document = self.db.json_collection.find_one({"_id": object_id})
        if not document:
            raise ValueError("JSON not found.")

        return {
            "document_id": str(document["_id"]),
            "file_name": document["file_name"],
            "json_data": document["json_data"],
        }

    def get_templates(self):
        template_folder = Path(Config.TEMPLATE_FOLDER)
        if not template_folder.exists():
            raise ValueError("Template folder not found.")

        return [{"template_name": file.stem,"file_name": file.name,}
            for file in sorted(template_folder.iterdir())
            if file.is_file() and file.suffix.lower() in {".xlsx", ".xlsm"}]

    def analyze_template(self, file):
        if not file or not file.filename:
            raise ValueError("Excel template file is required.")
        extension = Path(file.filename).suffix.lower()
        if extension not in {".xlsx", ".xlsm"}:
            raise ValueError("Only .xlsx and .xlsm templates are supported.")

        temp_path = None
        workbook = None
        try:
            with NamedTemporaryFile(suffix=extension,delete=False,) as temp_file:
                temp_path = Path(temp_file.name)
                file.save(temp_path)
            workbook = load_workbook(temp_path,data_only=False,keep_vba=(extension == ".xlsm"),)
            schema = self.helper.analyze_template_workbook(workbook,file_name=file.filename,)
            return schema

        except Exception as exc:
            logger.error(f"Template analysis failed: {exc}")
            raise

        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception:
                    pass

            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception as exc:
                    logger.warning(f"Temporary file cleanup failed: {exc}")

    def _build_and_store_mapping(self, document, template_path, sheet_name, workbook):
        excel_schema = self.helper.analyze_template_workbook(
            workbook, file_name=template_path.name, only_sheet=sheet_name,
        )
        json_schema = self.helper.analyze_json(document["json_data"])
        sheet_schema = excel_schema["sheets"][0]
        mappings = self.helper.build_mappings(
            json_schema=json_schema, excel_sheet_schema=sheet_schema,
        )
        if not mappings:
            raise ValueError("No JSON fields could be mapped to the Excel template.")
        mapping_doc = {
            "document_id": document["_id"],
            "template_file": template_path.name,
            "sheet_name": sheet_name,
            "mappings": mappings,
            "created_at": datetime.now(timezone.utc),
        }
        result = self.db.mapping_collection.insert_one(mapping_doc)
        logger.info(f"Mapping created: {result.inserted_id}")
        return str(result.inserted_id), mappings

    def create_mapping(self, document_id: str, template_file: str, sheet_name: str):
        workbook = None
        try:
            object_id = self.helper.to_object_id(document_id)
            document = self.db.json_collection.find_one({"_id": object_id})
            if not document:
                raise ValueError("JSON not found.")

            template_path = self.helper.safe_template_path(template_file)
            workbook = load_workbook( template_path, data_only=False, keep_vba=template_path.suffix.lower() == ".xlsm",)

            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist.")

            mapping_id, mappings = self._build_and_store_mapping(document, template_path, sheet_name, workbook)

            return {"mapping_id": mapping_id,"document_id": document_id, "template_file": template_path.name, 
                    "sheet_name": sheet_name, "mapping_count": len(mappings), "mappings": mappings,}

        except Exception as exc:
            logger.error(f"Mapping creation failed: {exc}")
            raise

        finally:
            if workbook:
                workbook.close()

    def generate_excel(self, data: dict[str, Any]):
        workbook = None
        try:
            document_id = data.get("document_id")
            template_file = data.get("template_file")
            sheet_name = data.get("sheet_name")
            mapping_id = data.get("mapping_id")
            document = self.db.json_collection.find_one({"_id": self.helper.to_object_id(document_id)})

            if not document:
                raise ValueError("JSON not found.")

            template_path = self.helper.safe_template_path(template_file)
            workbook = load_workbook(template_path,data_only=False, keep_vba=template_path.suffix.lower() == ".xlsm",)

            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' does not exist.")

            sheet = workbook[sheet_name]
            if mapping_id:
                mapping_document = self.db.mapping_collection.find_one({"_id": self.helper.to_object_id(mapping_id)})
                if not mapping_document:
                    raise ValueError("Mapping not found.")
                mappings = mapping_document["mappings"]
            else:
                mapping_id, mappings = self._build_and_store_mapping(document, template_path, sheet_name, workbook)
            write_result = self.helper.populate_template(workbook=workbook, sheet=sheet, json_data=document["json_data"], mappings=mappings,)
            generated_folder = Path(Config.GENERATED_FOLDER)
            generated_folder.mkdir(parents=True,exist_ok=True,)
            output_extension = (".xlsm"if template_path.suffix.lower() == ".xlsm" else ".xlsx")
            output_name = f"{__import__('uuid').uuid4().hex}{output_extension}"
            output_path = generated_folder / output_name
            workbook.save(output_path)
            export_document = {
                "document_id": self.helper.to_object_id(document_id),
                "mapping_id": (self.helper.to_object_id(mapping_id) if mapping_id else None),
                "template_file": template_path.name,
                "sheet_name": sheet_name,
                "file_name": output_name,
                "file_path": str(output_path),
                "write_result": write_result,
                "created_at": datetime.now(timezone.utc),
            }
            result = self.db.export_collection.insert_one(export_document)
            logger.info(f"Excel generated: {result.inserted_id}")

            return {
                "export_id": str(result.inserted_id),
                "file_name": output_name,
                "mapping_id": str(mapping_id) if mapping_id else None,
                "mapping_count": len(mappings),
                "written": write_result,
            }

        except Exception as exc:
            logger.error(f"Excel generation failed: {exc}")
            raise

        finally:
            if workbook:
                try:
                    workbook.close()
                except Exception as exc:
                    logger.error(f"Failed to close workbook: {exc}")

    def get_export_path(self, export_id: str):
        try:
            export = self.db.export_collection.find_one({"_id": self.helper.to_object_id(export_id)})
            if not export:
                raise ValueError("Export not found.")
            file_path = Path(export["file_path"])
            if not file_path.exists():
                raise ValueError("Generated Excel file no longer exists.")
            return str(file_path)
        except Exception as exc:
            logger.error(f"Failed to get export: {exc}")
            raise
